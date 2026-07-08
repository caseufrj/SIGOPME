from openpyxl import load_workbook

from database.database_service import DatabaseService


class ImportacaoExcelService:

    @staticmethod
    def importar_licitacoes(caminho_arquivo):

        workbook = load_workbook(
            caminho_arquivo,
            data_only=True
        )

        worksheet = workbook.active

        conn = DatabaseService.get_connection()

        cursor = conn.cursor()

        registros = 0

        for linha in worksheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            try:

                cursor.execute(
                    """
                    INSERT INTO Licitacoes (
                        NumeroLicitacao,
                        TipoLicitacao,
                        Ata,
                        Consignado,
                        Contrato,
                        NumeroContrato,
                        Processo,
                        Vigencia,
                        Fornecedor,
                        Cnpj,
                        CodConsig,
                        CodItem,
                        Sirep,
                        NomeMaterial,
                        QtdLicitada,
                        ValorUnd,
                        RP,
                        SC,
                        Especialidade
                    )
                    VALUES (
                        ?, ?, ?, ?, ?, ?,
                        ?, ?, ?, ?, ?, ?,
                        ?, ?, ?, ?, ?, ?, ?
                    )
                    """,
                    (
                        linha[0],
                        linha[1],
                        linha[2],
                        linha[3],
                        linha[4],
                        linha[5],
                        linha[6],
                        linha[7],
                        linha[8],
                        linha[9],
                        linha[10],
                        linha[11],
                        linha[12],
                        linha[13],
                        linha[14],
                        linha[15],
                        linha[24],
                        linha[25],
                        linha[26]
                    )
                )

                registros += 1

            except Exception as erro:
                print(f"Erro na linha: {erro}")

        conn.commit()
        conn.close()

        return registros
